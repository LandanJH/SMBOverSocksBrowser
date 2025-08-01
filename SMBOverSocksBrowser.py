import sys
import socket
import socks # PySocks library
import traceback
import io
import os
import ipaddress
import tempfile
import shutil
import subprocess
import json
from concurrent.futures import ThreadPoolExecutor, as_completed

# These libraries are for rendering specific file types.
import fitz # PyMuPDF for PDFs
import docx # python-docx for DOCX
import openpyxl # for XLSX

from PySide6.QtWidgets import (
    QApplication, QMainWindow, QWidget, QVBoxLayout, QGridLayout,
    QLabel, QLineEdit, QPushButton, QTreeView, QHeaderView, QMessageBox,
    QStatusBar, QComboBox, QHBoxLayout, QDialog, QTextEdit, QFileDialog,
    QScrollArea, QTabWidget, QMenu, QCheckBox, QStyle, QGroupBox
)
from PySide6.QtCore import QObject, Signal, Slot, QThread, QMetaObject, Qt, Q_ARG, QPoint, QProcess
from PySide6.QtGui import QStandardItemModel, QStandardItem, QPixmap, QFont, QIntValidator

# --- We now ONLY use the pysmb library for all SMB operations IN THE BROWSER ---
from smb.SMBConnection import SMBConnection
from smb import smb_structs
from smb.base import NotConnectedError, SMBTimeout
from smb.smb_structs import ProtocolError as SMBProtocolError


# This is the crucial step to make pysmb use the SOCKS proxy
smb_structs.socket = socks.socksocket

# Disable NTLMv1 support for better security
smb_structs.SUPPORT_NTLMv2 = True
smb_structs.SUPPORT_NTLMv1 = False

# --- NEW: Modern UI Stylesheets ---

# Modern Dark theme with red accents
MODERN_DARK_STYLESHEET = """
    /* General */
    QWidget {
        background-color: #2b2b2b;
        color: #f0f0f0;
        font-size: 14px;
        font-family: 'Segoe UI', 'Roboto', 'Helvetica Neue', 'Arial', sans-serif;
    }

    /* GroupBox */
    QGroupBox {
        font-weight: bold;
        border: 1px solid #4a4a4a;
        border-radius: 5px;
        margin-top: 10px;
    }
    QGroupBox::title {
        subcontrol-origin: margin;
        subcontrol-position: top center;
        padding: 0 10px;
    }

    /* Main Window & Dialogs */
    QMainWindow, QDialog, QScrollArea {
        background-color: #2b2b2b;
    }

    /* Labels & Status Bar */
    QLabel, QStatusBar {
        color: #cccccc;
    }

    /* TreeView */
    QTreeView {
        background-color: #3c3f41;
        border: 1px solid #4a4a4a;
        border-radius: 5px;
        alternate-background-color: #35383a;
        padding: 5px;
    }
    QTreeView::item {
        padding: 5px;
        border-radius: 3px;
    }
    QTreeView::item:selected {
        background-color: #E53935; /* Accent */
        color: #ffffff;
    }
    QHeaderView::section {
        background-color: #45494c;
        color: #f0f0f0;
        padding: 6px;
        border: none;
        border-bottom: 1px solid #2b2b2b;
    }

    /* Tab Widgets */
    QTabWidget::pane { 
        border: 1px solid #4a4a4a; 
        border-top: none;
        border-radius: 0 0 5px 5px;
    }
    QTabBar::tab { 
        background-color: #3c3f41; 
        color: #cccccc;
        padding: 10px 20px; 
        border: 1px solid #4a4a4a;
        border-bottom: none;
        border-top-left-radius: 5px;
        border-top-right-radius: 5px;
    }
    QTabBar::tab:selected { 
        background-color: #45494c; 
        color: #ffffff;
        border-color: #4a4a4a;
    }
    QTabBar::tab:!selected:hover {
        background-color: #414446;
    }

    /* Buttons */
    QPushButton {
        background-color: #555555;
        color: #f0f0f0;
        border: none;
        border-radius: 5px;
        padding: 8px 16px;
    }
    QPushButton:hover { 
        background-color: #6a6a6a; 
    }
    QPushButton:pressed {
        background-color: #757575;
    }
    QPushButton:disabled { 
        background-color: #404040; 
        color: #888888; 
    }
    
    /* Primary Action Buttons */
    QPushButton#PrimaryAction {
        background-color: #E53935; /* Accent */
        color: #ffffff;
        font-weight: bold;
    }
    QPushButton#PrimaryAction:hover {
        background-color: #C62828;
    }
    QPushButton#PrimaryAction:disabled {
        background-color: #404040;
        color: #888888;
    }

    /* Inputs */
    QLineEdit, QTextEdit, QComboBox {
        background-color: #3c3f41;
        border: 1px solid #4a4a4a;
        border-radius: 5px;
        padding: 6px;
    }
    QLineEdit:focus, QTextEdit:focus, QComboBox:focus {
        border: 1px solid #E53935; /* Accent on focus */
    }
    QComboBox::drop-down { 
        border: none;
    }

    /* Menu */
    QMenu {
        background-color: #3c3f41;
        border: 1px solid #4a4a4a;
    }
    QMenu::item:selected {
        background-color: #E53935; /* Accent */
        color: #ffffff;
    }

    /* CheckBox */
    QCheckBox::indicator {
        width: 16px;
        height: 16px;
        border: 1px solid #555;
        border-radius: 3px;
    }
    QCheckBox::indicator:checked {
        background-color: #E53935;
        border-color: #C62828;
    }
"""

# Modern Light theme with blue accents
MODERN_LIGHT_STYLESHEET = """
    /* General */
    QWidget {
        background-color: #f9f9f9;
        color: #333333;
        font-size: 14px;
        font-family: 'Segoe UI', 'Roboto', 'Helvetica Neue', 'Arial', sans-serif;
    }

    /* GroupBox */
    QGroupBox {
        font-weight: bold;
        border: 1px solid #e0e0e0;
        border-radius: 5px;
        margin-top: 10px;
    }
    QGroupBox::title {
        subcontrol-origin: margin;
        subcontrol-position: top center;
        padding: 0 10px;
    }

    /* Main Window & Dialogs */
    QMainWindow, QDialog, QScrollArea {
        background-color: #f9f9f9;
    }

    /* Labels & Status Bar */
    QLabel, QStatusBar {
        color: #444444;
    }

    /* TreeView */
    QTreeView {
        background-color: #ffffff;
        border: 1px solid #e0e0e0;
        border-radius: 5px;
        alternate-background-color: #f5f5f5;
        padding: 5px;
    }
    QTreeView::item {
        padding: 5px;
        border-radius: 3px;
    }
    QTreeView::item:selected {
        background-color: #1E88E5; /* Accent */
        color: #ffffff;
    }
    QHeaderView::section {
        background-color: #f0f0f0;
        color: #333333;
        padding: 6px;
        border: none;
        border-bottom: 1px solid #e0e0e0;
    }

    /* Tab Widgets */
    QTabWidget::pane { 
        border: 1px solid #e0e0e0; 
        border-top: none;
        border-radius: 0 0 5px 5px;
    }
    QTabBar::tab { 
        background-color: #f0f0f0; 
        color: #555555;
        padding: 10px 20px; 
        border: 1px solid #e0e0e0;
        border-bottom: none;
        border-top-left-radius: 5px;
        border-top-right-radius: 5px;
    }
    QTabBar::tab:selected { 
        background-color: #ffffff; 
        color: #333333;
        border-color: #e0e0e0;
    }
    QTabBar::tab:!selected:hover {
        background-color: #e7e7e7;
    }

    /* Buttons */
    QPushButton {
        background-color: #e0e0e0;
        color: #333333;
        border: none;
        border-radius: 5px;
        padding: 8px 16px;
    }
    QPushButton:hover { 
        background-color: #dcdcdc; 
    }
    QPushButton:pressed {
        background-color: #d0d0d0;
    }
    QPushButton:disabled { 
        background-color: #eeeeee; 
        color: #aaaaaa; 
    }

    /* Primary Action Buttons */
    QPushButton#PrimaryAction {
        background-color: #1E88E5; /* Accent */
        color: #ffffff;
        font-weight: bold;
    }
    QPushButton#PrimaryAction:hover {
        background-color: #1976D2;
    }
    QPushButton#PrimaryAction:disabled {
        background-color: #eeeeee;
        color: #aaaaaa;
    }

    /* Inputs */
    QLineEdit, QTextEdit, QComboBox {
        background-color: #ffffff;
        border: 1px solid #dcdcdc;
        border-radius: 5px;
        padding: 6px;
    }
    QLineEdit:focus, QTextEdit:focus, QComboBox:focus {
        border: 1px solid #1E88E5; /* Accent on focus */
    }
    QComboBox::drop-down { 
        border: none;
    }

    /* Menu */
    QMenu {
        background-color: #ffffff;
        border: 1px solid #e0e0e0;
    }
    QMenu::item:selected {
        background-color: #1E88E5; /* Accent */
        color: #ffffff;
    }

    /* CheckBox */
    QCheckBox::indicator {
        width: 16px;
        height: 16px;
        border: 1px solid #cccccc;
        border-radius: 3px;
    }
    QCheckBox::indicator:checked {
        background-color: #1E88E5;
        border-color: #1565C0;
    }
"""

class ProxyDialog(QDialog):
    """A dialog for adding or editing a SOCKS proxy entry."""
    def __init__(self, parent=None, name="", port=""):
        super().__init__(parent)
        self.setWindowTitle("Proxy Configuration")
        self.layout = QGridLayout(self)
        
        self.layout.addWidget(QLabel("Proxy Name:"), 0, 0)
        self.name_input = QLineEdit(name)
        self.layout.addWidget(self.name_input, 0, 1)

        self.layout.addWidget(QLabel("Port:"), 1, 0)
        self.port_input = QLineEdit(str(port))
        self.port_input.setValidator(QIntValidator(1, 65535, self))
        self.layout.addWidget(self.port_input, 1, 1)

        self.button_box = QHBoxLayout()
        self.ok_button = QPushButton("OK")
        self.ok_button.clicked.connect(self.accept)
        self.cancel_button = QPushButton("Cancel")
        self.cancel_button.clicked.connect(self.reject)
        self.button_box.addStretch()
        self.button_box.addWidget(self.ok_button)
        self.button_box.addWidget(self.cancel_button)

        self.layout.addLayout(self.button_box, 2, 0, 1, 2)

    def get_data(self):
        """Returns the entered data if the dialog is accepted."""
        if self.name_input.text() and self.port_input.text():
            return self.name_input.text().strip(), int(self.port_input.text())
        return None, None

class ImageLabel(QLabel):
    def __init__(self, pixmap: QPixmap, parent=None):
        super().__init__(parent)
        self._pixmap = pixmap
        self.setMinimumSize(1, 1)
        self.setAlignment(Qt.AlignCenter)
        self.setPixmap(self._pixmap.scaled(self.size(), Qt.KeepAspectRatio, Qt.SmoothTransformation))
    def resizeEvent(self, event):
        self.setPixmap(self._pixmap.scaled(self.size(), Qt.KeepAspectRatio, Qt.SmoothTransformation))
        super().resizeEvent(event)

class ConversionWorker(QObject):
    conversion_successful = Signal(bytes)
    conversion_failed = Signal(str)
    def __init__(self, content_bytes, extension, lo_path):
        super().__init__()
        self.content_bytes = content_bytes; self.extension = extension; self.lo_path = lo_path
    @Slot()
    def do_conversion(self):
        temp_dir = tempfile.mkdtemp()
        try:
            input_path = os.path.join(temp_dir, f"tempfile{self.extension}")
            with open(input_path, 'wb') as f: f.write(self.content_bytes)
            cmd = [self.lo_path, '--headless', '--convert-to', 'pdf', '--outdir', temp_dir, input_path]
            subprocess.run(cmd, timeout=30, check=True, capture_output=True)
            output_pdf_path = os.path.join(temp_dir, "tempfile.pdf")
            if os.path.exists(output_pdf_path):
                with open(output_pdf_path, 'rb') as f: pdf_bytes = f.read()
                self.conversion_successful.emit(pdf_bytes)
            else: raise FileNotFoundError("LibreOffice did not create the PDF file.")
        except subprocess.TimeoutExpired: self.conversion_failed.emit("LibreOffice conversion timed out after 30 seconds.")
        except subprocess.CalledProcessError as e:
            error_output = e.stderr.decode('utf-8', errors='ignore')
            self.conversion_failed.emit(f"LibreOffice failed to convert document.\n\nError:\n{error_output}")
        except Exception as e: self.conversion_failed.emit(f"Failed to render file using LibreOffice:\n{e}\n\nMake sure LibreOffice is installed and accessible.")
        finally: shutil.rmtree(temp_dir)

class PreviewDialog(QDialog):
    def __init__(self, file_name, content_bytes, parent=None):
        super().__init__(parent)
        self.setWindowTitle(f"Preview: {file_name}"); self.setGeometry(150, 150, 800, 600)
        self.main_layout = QVBoxLayout(self)
        file_ext = os.path.splitext(file_name)[1].lower()
        office_formats = ['.doc', '.docx', '.xls', '.xlsx']; image_formats = ['.png', '.jpg', '.jpeg', '.gif', '.bmp']
        libreoffice_path = self._find_libreoffice_path()
        if file_ext in office_formats and libreoffice_path:
            self.loading_label = QLabel("Converting document, please wait...", self)
            self.loading_label.setAlignment(Qt.AlignCenter)
            self.main_layout.addWidget(self.loading_label)
            self.start_libreoffice_conversion(content_bytes, file_ext, libreoffice_path)
        elif file_ext in image_formats: self.setup_image_preview(content_bytes, self.main_layout)
        elif file_ext == '.pdf': self.setup_pdf_preview(content_bytes, self.main_layout)
        elif file_ext == '.docx': self.setup_docx_preview(content_bytes, self.main_layout)
        elif file_ext == '.xlsx': self.setup_xlsx_preview(content_bytes, self.main_layout)
        else: self.setup_text_preview(content_bytes, self.main_layout)
    def _find_libreoffice_path(self):
        if sys.platform == "win32":
            paths = [os.path.join(os.environ["ProgramFiles"], "LibreOffice", "program", "soffice.exe"), os.path.join(os.environ["ProgramFiles(x86)"], "LibreOffice", "program", "soffice.exe")]
        elif sys.platform == "darwin": paths = ["/Applications/LibreOffice.app/Contents/MacOS/soffice"]
        else: paths = ["/usr/bin/soffice", "/usr/lib/libreoffice/program/soffice"]
        for path in paths:
            if os.path.exists(path): return path
        return None
    def start_libreoffice_conversion(self, content_bytes, extension, lo_path):
        self.thread = QThread(); self.worker = ConversionWorker(content_bytes, extension, lo_path)
        self.worker.moveToThread(self.thread); self.worker.conversion_successful.connect(self.on_pdf_ready)
        self.worker.conversion_failed.connect(self.on_conversion_error); self.thread.started.connect(self.worker.do_conversion)
        self.worker.conversion_successful.connect(self.thread.quit); self.worker.conversion_failed.connect(self.thread.quit)
        self.thread.finished.connect(self.thread.deleteLater); self.worker.moveToThread(self.thread); self.thread.start()
    @Slot(bytes)
    def on_pdf_ready(self, pdf_bytes):
        self.loading_label.hide(); self.loading_label.deleteLater(); self.setup_pdf_preview(pdf_bytes, self.main_layout)
    @Slot(str)
    def on_conversion_error(self, error_msg): self.loading_label.setText(error_msg)
    def setup_text_preview(self, content_bytes, layout):
        text_edit = QTextEdit(); text_edit.setReadOnly(True)
        try: text_content = content_bytes.decode('utf-8')
        except UnicodeDecodeError: text_content = content_bytes.decode('latin-1', errors='replace')
        text_edit.setPlainText(text_content); layout.addWidget(text_edit)
    def setup_image_preview(self, content_bytes, layout):
        pixmap = QPixmap(); pixmap.loadFromData(content_bytes)
        image_label = ImageLabel(pixmap); scroll_area = QScrollArea()
        scroll_area.setWidgetResizable(True); scroll_area.setWidget(image_label); layout.addWidget(scroll_area)
    def setup_pdf_preview(self, content_bytes, layout):
        scroll_area = QScrollArea(); scroll_area.setWidgetResizable(True)
        container = QWidget(); pdf_layout = QVBoxLayout(container)
        try:
            doc = fitz.open(stream=content_bytes, filetype="pdf")
            for page_num in range(len(doc)):
                page = doc.load_page(page_num); pix = page.get_pixmap()
                img = QPixmap(); img.loadFromData(pix.tobytes()); label = QLabel(); label.setPixmap(img)
                pdf_layout.addWidget(label)
        except Exception as e: pdf_layout.addWidget(QLabel(f"Error rendering PDF: {e}"))
        scroll_area.setWidget(container); layout.addWidget(scroll_area)
    def setup_docx_preview(self, content_bytes, layout):
        text_edit = QTextEdit(); text_edit.setReadOnly(True)
        try:
            doc_stream = io.BytesIO(content_bytes); doc = docx.Document(doc_stream)
            full_text = [para.text for para in doc.paragraphs]
            text_edit.setPlainText('\n'.join(full_text))
        except Exception as e: text_edit.setPlainText(f"Error rendering DOCX: {e}")
        layout.addWidget(text_edit)
    def setup_xlsx_preview(self, content_bytes, layout):
        text_edit = QTextEdit(); text_edit.setReadOnly(True); text_edit.setFont(QFont("Courier New", 10))
        try:
            wb_stream = io.BytesIO(content_bytes); wb = openpyxl.load_workbook(wb_stream, read_only=True)
            sheet = wb.active; max_widths = {}
            for row in sheet.iter_rows():
                for i, cell in enumerate(row):
                    max_widths.setdefault(i, 0)
                    if cell.value: max_widths[i] = max(max_widths[i], len(str(cell.value)))
            table_text = ""
            for row in sheet.iter_rows():
                row_text = []
                for i, cell in enumerate(row):
                    val = str(cell.value) if cell.value is not None else ""
                    row_text.append(val.ljust(max_widths.get(i, 0)))
                table_text += " | ".join(row_text) + "\n"
            text_edit.setPlainText(table_text)
        except Exception as e: text_edit.setPlainText(f"Error rendering XLSX: {e}")
        layout.addWidget(text_edit)

class BrowserWorker(QObject):
    """Worker for browser tab with search caching."""
    connection_success = Signal(list); connection_failed = Signal(str)
    status_update = Signal(str); preview_content_ready = Signal(str, bytes)
    preview_failed = Signal(str); download_finished = Signal(str)
    download_failed = Signal(str); search_finished = Signal(list)
    
    def __init__(self, config):
        super().__init__()
        self.config = config
        self.smb_connection = None
        self._is_running = True
        self.is_cached = False
        self.is_caching = False # Flag to prevent concurrent caching
        self.file_path_cache = []
        
    @Slot()
    def run_browser(self):
        try:
            if not self._is_running: return
            self.status_update.emit("Setting SOCKS proxy...")
            if self.config.get('use_proxy', True):
                socks.set_default_proxy(socks.SOCKS5, self.config['proxy_host'], self.config['proxy_port'])
            else:
                socks.set_default_proxy(None)
            socket.socket = socks.socksocket
            
            self.status_update.emit(f"Connecting to SMB host {self.config['smb_host']}...")
            self.smb_connection = SMBConnection(self.config['smb_user'], self.config['smb_pass'], "pyside-smb-browser", "remote-server", use_ntlm_v2=True, is_direct_tcp=True)
            if not self.smb_connection.connect(self.config['smb_host'], 445): raise Exception("Connection returned False")
            self.connection_success.emit(self.browse_path("/"))
        except Exception as e:
            if self._is_running: self.connection_failed.emit(str(e))

    @Slot()
    def start_background_caching(self):
        """Starts the file indexing process in the background."""
        if not self.smb_connection or self.is_cached or self.is_caching or not self._is_running:
            return
        try:
            self.is_caching = True
            self.status_update.emit("Starting background file indexing...")
            self.file_path_cache = []  # Clear previous cache
            self._build_cache_parallel()
            self.is_cached = True
            self.status_update.emit("Background file indexing is complete. Search is now available.")
        except Exception as e:
            self.status_update.emit(f"Error during background caching: {e}")
        finally:
            self.is_caching = False # Ensure this is always reset
    
    @Slot(str)
    def do_search(self, keyword):
        if not self.smb_connection: return

        if self.is_caching:
            self.status_update.emit("File indexing is in progress. Please wait a moment before searching.")
            self.search_finished.emit([]) # Return an empty result to avoid user confusion
            return
            
        try:
            if not self.is_cached:
                self.is_caching = True
                self.status_update.emit("First search: building file index... This may take a moment.")
                self.file_path_cache = []
                self._build_cache_parallel()
                self.is_cached = True
                self.is_caching = False

            self.status_update.emit(f"Searching for '{keyword}' in cache...")
            # Search results will now be a list of tuples (path, size)
            results = [(path, size) for path, size in self.file_path_cache if keyword.lower() in os.path.basename(path).lower()]
            self.search_finished.emit(results)
            
        except Exception as e:
            self.status_update.emit(f"Error during search: {e}")
            self.search_finished.emit([])
        finally:
            # Ensure reset in case of an error during the initial search caching
            if self.is_caching:
                self.is_caching = False

    def _build_cache_parallel(self):
        dirs_to_scan = ['/']
        with ThreadPoolExecutor(max_workers=20) as executor:
            while dirs_to_scan:
                if not self._is_running: break
                next_level_dirs = []
                future_to_path = {executor.submit(self._fetch_path_contents, path): path for path in dirs_to_scan}
                
                for future in as_completed(future_to_path):
                    if not self._is_running: break
                    try:
                        files, subdirs = future.result()
                        self.file_path_cache.extend(files)
                        next_level_dirs.extend(subdirs)
                    except Exception as e:
                        print(f"Error fetching path contents during cache build: {e}")
                dirs_to_scan = next_level_dirs

    def _fetch_path_contents(self, path):
        if not self._is_running: return [], []
        self.status_update.emit(f"Indexing: {path}")
        
        if self.config.get('use_proxy', True):
            socks.set_default_proxy(socks.SOCKS5, self.config['proxy_host'], self.config['proxy_port'])
        else:
            socks.set_default_proxy(None)
        socket.socket = socks.socksocket

        conn = SMBConnection(self.config['smb_user'], self.config['smb_pass'], f"cache-builder-{os.urandom(4).hex()}", "remote-server", use_ntlm_v2=True, is_direct_tcp=True)
        conn.connect(self.config['smb_host'], 445)

        files_found = []; subdirs_found = []
        for item in conn.listPath(self.config['smb_share'], path):
            if item.filename in ['.', '..']: continue
            full_path = os.path.join(path, item.filename).replace('\\', '/')
            if item.isDirectory:
                subdirs_found.append(full_path)
            else:
                files_found.append((full_path, item.file_size))
        conn.close()
        return files_found, subdirs_found
        
    @Slot(str)
    def do_preview(self, remote_path):
        try:
            if not self._is_running or not self.smb_connection: return
            file_obj = io.BytesIO(); file_name = os.path.basename(remote_path)
            self.smb_connection.retrieveFile(self.config['smb_share'], remote_path, file_obj)
            self.preview_content_ready.emit(file_name, file_obj.getvalue())
        except Exception as e: self.preview_failed.emit(str(e))

    @Slot(str, str)
    def do_download(self, remote_path, local_path):
        try:
            if not self._is_running or not self.smb_connection: return
            with open(local_path, 'wb') as f: self.smb_connection.retrieveFile(self.config['smb_share'], remote_path, f)
            self.download_finished.emit(local_path)
        except Exception as e: self.download_failed.emit(str(e))

    def browse_path(self, path):
        if not self.smb_connection: return []
        results = []
        files = self.smb_connection.listPath(self.config['smb_share'], path)
        dirs = sorted([f.filename for f in files if f.isDirectory and f.filename not in ['.', '..']])
        regular_files = sorted([f for f in files if not f.isDirectory], key=lambda x: x.filename)
        for d in dirs: results.append({'name': d, 'type': 'dir', 'size': 0})
        for f in regular_files: results.append({'name': f.filename, 'type': 'file', 'size': f.file_size})
        return results

    @Slot()
    def stop(self):
        self._is_running = False
        if self.smb_connection:
            try: self.smb_connection.close()
            except Exception: pass

class SMBBrowserApp(QMainWindow):
    def __init__(self):
        super().__init__()
        self.setWindowTitle("SMB over SOCKS Proxy Browser")
        self.setGeometry(100, 100, 900, 700)
        self.config = {}
        self.load_config()
        self.browser_worker_thread = None; self.browser_worker = None
        self.scanner_process = None
        self.scanner_buffer = ""
        self.current_smb_path = "/"; self.is_in_search_mode = False
        self.create_widgets()
        self.apply_theme(self.config.get('theme', 'dark'))


    def load_config(self):
        self.config_file = "config.json"
        default_config = {
            "proxies": {"example1": 1337, "example2": 1338, "example3": 1339},
            "theme": "dark",
            "auto_index": True
        }
        if os.path.exists(self.config_file):
            try:
                with open(self.config_file, 'r') as f:
                    loaded_config = json.load(f)

                # Migration logic for old config format
                if "proxies" not in loaded_config and "theme" not in loaded_config and loaded_config:
                    self.config = { "proxies": loaded_config, "theme": "dark", "auto_index": True }
                    self.save_config()
                else:
                    self.config = loaded_config
                
                # Ensure all default keys are present
                for key, value in default_config.items():
                    if key not in self.config:
                        self.config[key] = value

            except (json.JSONDecodeError, IOError) as e:
                QMessageBox.warning(self, "Config Error", f"Could not load '{self.config_file}':\n{e}\n\nFalling back to default settings.")
                self.config = default_config
                self.save_config()
        else:
            self.config = default_config
            self.save_config()

    def save_config(self):
        try:
            with open(self.config_file, 'w') as f:
                json.dump(self.config, f, indent=4)
        except IOError as e:
            print(f"Could not save config file '{self.config_file}':\n{e}")

    def create_widgets(self):
        central_widget = QWidget(); self.setCentralWidget(central_widget)
        main_layout = QVBoxLayout(central_widget); self.tabs = QTabWidget()
        main_layout.setSpacing(10)
        main_layout.setContentsMargins(10, 10, 10, 10)

        main_layout.addWidget(self.tabs)
        self.browser_tab = QWidget()
        self.scanner_tab = QWidget()
        self.settings_tab = QWidget()
        self.tabs.addTab(self.browser_tab, "Share Browser")
        self.tabs.addTab(self.scanner_tab, "Subnet Scanner")
        self.tabs.addTab(self.settings_tab, "Settings")

        self.create_browser_tab()
        self.create_scanner_tab()
        self.create_settings_tab() # New settings tab

        self.setStatusBar(QStatusBar(self))
        self.statusBar().showMessage("Status: Disconnected")

    def create_browser_tab(self):
        layout = QVBoxLayout(self.browser_tab); grid_layout = QGridLayout()
        layout.setSpacing(10)
        grid_layout.setSpacing(10)
        
        grid_layout.addWidget(QLabel("SOCKS Proxy:"), 0, 0); self.proxy_selector = QComboBox()
        grid_layout.addWidget(self.proxy_selector, 0, 1, 1, 3)
        grid_layout.addWidget(QLabel("Remote SMB Host:"), 1, 0); self.smb_host = QLineEdit("192.168.1.100")
        grid_layout.addWidget(self.smb_host, 1, 1); grid_layout.addWidget(QLabel("SMB Share Name:"), 1, 2)
        self.smb_share = QLineEdit("SharedFolder"); grid_layout.addWidget(self.smb_share, 1, 3)
        grid_layout.addWidget(QLabel("SMB User:"), 2, 0); self.smb_user = QLineEdit("")
        grid_layout.addWidget(self.smb_user, 2, 1); grid_layout.addWidget(QLabel("SMB Password:"), 2, 2)
        self.smb_pass = QLineEdit(""); self.smb_pass.setEchoMode(QLineEdit.Password)
        grid_layout.addWidget(self.smb_pass, 2, 3); layout.addLayout(grid_layout)

        search_layout = QHBoxLayout(); self.search_input = QLineEdit()
        search_layout.setSpacing(10)
        self.search_input.setPlaceholderText("Enter keyword to search for files in the current share (builds cache on first run)..."); self.search_button = QPushButton("Search")
        self.cancel_search_button = QPushButton("Cancel Search"); self.clear_search_button = QPushButton("Clear Search")
        search_layout.addWidget(self.search_input); search_layout.addWidget(self.search_button)
        search_layout.addWidget(self.cancel_search_button); search_layout.addWidget(self.clear_search_button)
        layout.addLayout(search_layout); action_layout = QHBoxLayout()
        action_layout.setSpacing(10)

        self.connect_button = QPushButton("Connect")
        self.connect_button.setObjectName("PrimaryAction")
        self.disconnect_button = QPushButton("Disconnect")
        self.disconnect_button.setEnabled(False); self.preview_button = QPushButton("Preview")
        self.preview_button.setEnabled(False); self.download_button = QPushButton("Download")
        self.download_button.setEnabled(False); action_layout.addWidget(self.connect_button)
        action_layout.addWidget(self.disconnect_button); action_layout.addStretch()
        action_layout.addWidget(self.preview_button); action_layout.addWidget(self.download_button)
        layout.addLayout(action_layout)

        self.search_button.clicked.connect(self.start_search)
        self.cancel_search_button.clicked.connect(self.stop_search); self.clear_search_button.clicked.connect(self.clear_search)
        self.search_button.setEnabled(False); self.cancel_search_button.setVisible(False); self.clear_search_button.setEnabled(False)
        self.path_label = QLabel("Current Path: /"); layout.addWidget(self.path_label)
        self.file_tree = QTreeView()
        self.file_tree.setAlternatingRowColors(True)
        self.file_tree.doubleClicked.connect(self.on_item_double_clicked); layout.addWidget(self.file_tree)
        self.model = QStandardItemModel(); self.model.setHorizontalHeaderLabels(['Name', 'Size'])
        self.file_tree.setModel(self.model); self.file_tree.selectionModel().selectionChanged.connect(self.on_selection_changed)
        self.connect_button.clicked.connect(self.start_connection); self.disconnect_button.clicked.connect(self.disconnect)
        self.preview_button.clicked.connect(self.start_preview); self.download_button.clicked.connect(self.start_download)
        
        self.file_tree.header().setSectionResizeMode(0, QHeaderView.Stretch)
        self.file_tree.header().setSectionResizeMode(1, QHeaderView.ResizeToContents)

    def create_scanner_tab(self):
        layout = QVBoxLayout(self.scanner_tab)
        layout.setSpacing(10)
        grid = QGridLayout()
        grid.setSpacing(10)
        
        grid.addWidget(QLabel("SOCKS Proxy:"), 0, 0); self.scan_proxy_selector = QComboBox()
        grid.addWidget(self.scan_proxy_selector, 0, 1)
        grid.addWidget(QLabel("Subnet (CIDR):"), 1, 0); self.scan_subnet_input = QLineEdit("192.168.1.0/24")
        grid.addWidget(self.scan_subnet_input, 1, 1); grid.addWidget(QLabel("Username:"), 2, 0)
        self.scan_user_input = QLineEdit("Guest"); grid.addWidget(self.scan_user_input, 2, 1)
        grid.addWidget(QLabel("Password:"), 3, 0); self.scan_pass_input = QLineEdit("")
        self.scan_pass_input.setEchoMode(QLineEdit.Password); grid.addWidget(self.scan_pass_input, 3, 1)
        layout.addLayout(grid); self.quick_scan_checkbox = QCheckBox("Quick Scan (Skips slow permission checks, may be inconsistent)")
        self.quick_scan_checkbox.setChecked(True); layout.addWidget(self.quick_scan_checkbox)
        
        scan_action_layout = QHBoxLayout()
        scan_action_layout.setSpacing(10)
        self.start_scan_button = QPushButton("Start Scan")
        self.start_scan_button.setObjectName("PrimaryAction")
        self.cancel_scan_button = QPushButton("Cancel Scan"); self.open_share_button = QPushButton("Open in Browser")
        self.cancel_scan_button.setVisible(False); self.open_share_button.setEnabled(False)
        scan_action_layout.addWidget(self.start_scan_button); scan_action_layout.addWidget(self.cancel_scan_button)
        scan_action_layout.addStretch(); scan_action_layout.addWidget(self.open_share_button)
        layout.addLayout(scan_action_layout); self.scan_results_tree = QTreeView()
        self.scan_results_tree.header().setSectionResizeMode(QHeaderView.Stretch)
        self.scan_results_tree.setContextMenuPolicy(Qt.CustomContextMenu)
        self.scan_results_tree.customContextMenuRequested.connect(self.show_scanner_context_menu)
        layout.addWidget(self.scan_results_tree); self.scan_model = QStandardItemModel()
        self.scan_model.setHorizontalHeaderLabels(['Host', 'Share', 'Permissions']); self.scan_results_tree.setModel(self.scan_model)
        self.start_scan_button.clicked.connect(self.start_scan); self.cancel_scan_button.clicked.connect(self.stop_scan)
        self.open_share_button.clicked.connect(self.open_share_in_browser)
        self.scan_results_tree.selectionModel().selectionChanged.connect(self.on_scanner_selection_changed)
        
        # Populate proxy selectors now that all tabs are created
        self.update_proxy_selectors()
        
    def create_settings_tab(self):
        layout = QVBoxLayout(self.settings_tab)
        layout.setSpacing(15)
        layout.setContentsMargins(20, 20, 20, 20)

        # --- Proxy Management ---
        proxy_group = QGroupBox("Proxy Management")
        proxy_layout = QVBoxLayout()
        
        self.proxy_settings_model = QStandardItemModel()
        self.proxy_settings_model.setHorizontalHeaderLabels(["Name", "Port"])
        self.proxy_settings_tree = QTreeView()
        self.proxy_settings_tree.setModel(self.proxy_settings_model)
        self.proxy_settings_tree.setAlternatingRowColors(True)
        self.proxy_settings_tree.header().setSectionResizeMode(0, QHeaderView.Stretch)
        proxy_layout.addWidget(self.proxy_settings_tree)

        proxy_buttons = QHBoxLayout()
        add_proxy_btn = QPushButton("Add")
        edit_proxy_btn = QPushButton("Edit")
        remove_proxy_btn = QPushButton("Remove")
        proxy_buttons.addStretch()
        proxy_buttons.addWidget(add_proxy_btn)
        proxy_buttons.addWidget(edit_proxy_btn)
        proxy_buttons.addWidget(remove_proxy_btn)
        proxy_layout.addLayout(proxy_buttons)
        
        proxy_group.setLayout(proxy_layout)
        layout.addWidget(proxy_group)

        # --- Appearance Settings ---
        appearance_group = QGroupBox("Appearance")
        appearance_layout = QGridLayout()
        appearance_layout.addWidget(QLabel("Theme:"), 0, 0)
        self.theme_selector = QComboBox()
        self.theme_selector.addItems(["Dark", "Light"])
        self.theme_selector.setCurrentText(self.config.get('theme', 'dark').capitalize())
        appearance_layout.addWidget(self.theme_selector, 0, 1)
        appearance_group.setLayout(appearance_layout)
        layout.addWidget(appearance_group)

        # --- Behavior Settings ---
        behavior_group = QGroupBox("Behavior")
        behavior_layout = QVBoxLayout()
        self.auto_index_checkbox = QCheckBox("Automatically index share contents on connect")
        self.auto_index_checkbox.setChecked(self.config.get('auto_index', True))
        behavior_layout.addWidget(self.auto_index_checkbox)
        behavior_group.setLayout(behavior_layout)
        layout.addWidget(behavior_group)

        layout.addStretch() # Push all settings to the top

        # --- Connect signals ---
        add_proxy_btn.clicked.connect(self.add_proxy)
        edit_proxy_btn.clicked.connect(self.edit_proxy)
        remove_proxy_btn.clicked.connect(self.remove_proxy)
        self.theme_selector.currentTextChanged.connect(self.update_theme_setting)
        self.auto_index_checkbox.toggled.connect(self.update_auto_index_setting)
        
        # Populate proxy list
        self.populate_proxy_settings_list()

    def populate_proxy_settings_list(self):
        """Refreshes the list of proxies in the settings tab."""
        self.proxy_settings_model.clear()
        self.proxy_settings_model.setHorizontalHeaderLabels(["Name", "Port"])
        proxies = self.config.get('proxies', {})
        for name, port in sorted(proxies.items()):
            name_item = QStandardItem(name)
            port_item = QStandardItem(str(port))
            self.proxy_settings_model.appendRow([name_item, port_item])

    def update_proxy_selectors(self):
        """Updates the proxy dropdowns in the browser and scanner tabs."""
        # Store current selections
        current_browser_proxy = self.proxy_selector.currentText()
        current_scanner_proxy = self.scan_proxy_selector.currentText()

        # Clear and repopulate
        self.proxy_selector.clear()
        self.scan_proxy_selector.clear()
        self.proxy_selector.addItem("None (Direct Scan)")
        self.scan_proxy_selector.addItem("None (Direct Scan)")
        
        proxy_names = sorted(self.config.get('proxies', {}).keys())
        self.proxy_selector.addItems(proxy_names)
        self.scan_proxy_selector.addItems(proxy_names)

        # Restore selections if they still exist
        if self.proxy_selector.findText(current_browser_proxy) != -1:
            self.proxy_selector.setCurrentText(current_browser_proxy)
        if self.scan_proxy_selector.findText(current_scanner_proxy) != -1:
            self.scan_proxy_selector.setCurrentText(current_scanner_proxy)

    @Slot()
    def add_proxy(self):
        dialog = ProxyDialog(self)
        if dialog.exec() == QDialog.Accepted:
            name, port = dialog.get_data()
            if name and port:
                if name in self.config['proxies']:
                    QMessageBox.warning(self, "Error", f"A proxy with the name '{name}' already exists.")
                    return
                self.config['proxies'][name] = port
                self.populate_proxy_settings_list()
                self.update_proxy_selectors()

    @Slot()
    def edit_proxy(self):
        indexes = self.proxy_settings_tree.selectionModel().selectedIndexes()
        if not indexes:
            QMessageBox.information(self, "Edit Proxy", "Please select a proxy to edit.")
            return
        
        row = indexes[0].row()
        old_name = self.proxy_settings_model.item(row, 0).text()
        old_port = self.proxy_settings_model.item(row, 1).text()

        dialog = ProxyDialog(self, name=old_name, port=old_port)
        if dialog.exec() == QDialog.Accepted:
            new_name, new_port = dialog.get_data()
            if new_name and new_port:
                # Remove old entry
                del self.config['proxies'][old_name]
                # Add new entry
                self.config['proxies'][new_name] = new_port
                self.populate_proxy_settings_list()
                self.update_proxy_selectors()
    
    @Slot()
    def remove_proxy(self):
        indexes = self.proxy_settings_tree.selectionModel().selectedIndexes()
        if not indexes:
            QMessageBox.information(self, "Remove Proxy", "Please select a proxy to remove.")
            return

        row = indexes[0].row()
        name_to_remove = self.proxy_settings_model.item(row, 0).text()

        reply = QMessageBox.question(self, "Confirm Removal", f"Are you sure you want to remove the proxy '{name_to_remove}'?", QMessageBox.Yes | QMessageBox.No, QMessageBox.No)
        if reply == QMessageBox.Yes:
            if name_to_remove in self.config['proxies']:
                del self.config['proxies'][name_to_remove]
                self.populate_proxy_settings_list()
                self.update_proxy_selectors()

    @Slot(str)
    def update_theme_setting(self, theme_text):
        theme = theme_text.lower()
        self.apply_theme(theme)

    @Slot(bool)
    def update_auto_index_setting(self, checked):
        self.config['auto_index'] = checked
    
    def apply_theme(self, theme_name):
        app = QApplication.instance()
        if theme_name == 'light':
            app.setStyleSheet(MODERN_LIGHT_STYLESHEET)
        else: # Default to dark
            app.setStyleSheet(MODERN_DARK_STYLESHEET)
        self.config['theme'] = theme_name
        # Update selector in settings if it exists
        if hasattr(self, 'theme_selector'):
            self.theme_selector.setCurrentText(theme_name.capitalize())

    def get_icon_for_filename(self, filename):
        style = QApplication.style()
        extension = os.path.splitext(filename)[1].lower()
        
        icon_map = {
            '.zip': QStyle.SP_DriveHDIcon, '.rar': QStyle.SP_DriveHDIcon, '.7z': QStyle.SP_DriveHDIcon,
            '.tar': QStyle.SP_DriveHDIcon, '.gz': QStyle.SP_DriveHDIcon,
            '.png': QStyle.SP_FileDialogDetailedView, '.jpg': QStyle.SP_FileDialogDetailedView, '.jpeg': QStyle.SP_FileDialogDetailedView,
            '.gif': QStyle.SP_FileDialogDetailedView, '.bmp': QStyle.SP_FileDialogDetailedView,
            '.db': QStyle.SP_DriveHDIcon, '.sqlite': QStyle.SP_DriveHDIcon, '.sqlite3': QStyle.SP_DriveHDIcon,
            '.mdb': QStyle.SP_DriveHDIcon, '.accdb': QStyle.SP_DriveHDIcon,
        }
        
        icon_enum = icon_map.get(extension, QStyle.SP_FileIcon)
        return style.standardIcon(icon_enum)

    def format_file_size(self, size_bytes):
        if size_bytes is None or not isinstance(size_bytes, (int, float)) or size_bytes <= 0:
            return "" # Show nothing for 0 bytes or directories
        power = 1024
        n = 0
        power_labels = {0: 'B', 1: 'KB', 2: 'MB', 3: 'GB', 4: 'TB'}
        while size_bytes >= power and n < len(power_labels) -1:
            size_bytes /= power
            n += 1
        return f"{round(size_bytes)} {power_labels.get(n, 'B')}"


    @Slot()
    def start_connection(self):
        self.disconnect()
        proxy_selection = self.proxy_selector.currentText()
        config = {'smb_host': self.smb_host.text(), 'smb_share': self.smb_share.text(), 'smb_user': self.smb_user.text(), 'smb_pass': self.smb_pass.text()}
        if proxy_selection == "None (Direct Scan)": config['use_proxy'] = False
        else:
            if not proxy_selection or proxy_selection not in self.config['proxies']:
                 QMessageBox.warning(self, "Connection Error", f"Proxy '{proxy_selection}' not found in settings."); return
            config['use_proxy'] = True; config['proxy_host'] = '127.0.0.1'; config['proxy_port'] = self.config['proxies'][proxy_selection]
        self.connect_button.setEnabled(False); self.proxy_selector.setEnabled(False)
        self.browser_worker_thread = QThread(); self.browser_worker = BrowserWorker(config)
        self.browser_worker.moveToThread(self.browser_worker_thread); self.browser_worker.status_update.connect(self.update_status)
        self.browser_worker.connection_success.connect(self.on_connection_success); self.browser_worker.connection_failed.connect(self.on_connection_failed)
        self.browser_worker.preview_content_ready.connect(self.on_preview_ready)
        self.browser_worker.preview_failed.connect(lambda e: QMessageBox.warning(self, "Preview Failed", str(e)))
        self.browser_worker.download_finished.connect(lambda p: QMessageBox.information(self, "Success", f"File downloaded to:\n{p}"))
        self.browser_worker.download_failed.connect(lambda e: QMessageBox.warning(self, "Download Failed", str(e)))
        self.browser_worker.search_finished.connect(self.on_search_finished)
        self.browser_worker_thread.started.connect(self.browser_worker.run_browser); self.browser_worker_thread.start()

    @Slot()
    def start_search(self):
        keyword = self.search_input.text()
        if not keyword: QMessageBox.information(self, "Search", "Please enter a keyword."); return
        if self.browser_worker:
            self.search_button.setEnabled(False); self.cancel_search_button.setEnabled(False)
            self.clear_search_button.setEnabled(False); self.is_in_search_mode = True
            QMetaObject.invokeMethod(self.browser_worker, 'do_search', Qt.QueuedConnection, Q_ARG(str, keyword))
    @Slot()
    def stop_search(self): pass # The cancel button is now effectively disabled during search
    @Slot()
    def clear_search(self):
        self.is_in_search_mode = False; self.search_input.clear(); self.browse_path(self.current_smb_path)
    
    @Slot(list)
    def on_search_finished(self, results):
        self.model.clear(); self.model.setHorizontalHeaderLabels(['Search Result (Full Path)', 'Size'])
        if not results and self.browser_worker and self.browser_worker.is_caching:
            # Don't show "no results" if the message is just because caching is in progress
            return
        if not results:
             self.model.appendRow(QStandardItem("No matching files found."))
        else:
            for path, size in results:
                icon = self.get_icon_for_filename(path)
                item = QStandardItem(icon, path)
                item.setData("file", Qt.UserRole) # Mark as file for double click
                size_str = self.format_file_size(size)
                size_item = QStandardItem(size_str)
                size_item.setTextAlignment(Qt.AlignRight | Qt.AlignVCenter)
                size_item.setEditable(False)
                item.setEditable(False)
                self.model.appendRow([item, size_item])
        self.path_label.setText(f"Found {len(results)} results for '{self.search_input.text()}'")
        self.search_button.setEnabled(True); self.cancel_search_button.setEnabled(False); self.clear_search_button.setEnabled(True)
        # Adjust column sizes for search results
        self.file_tree.header().setSectionResizeMode(0, QHeaderView.Stretch)
        self.file_tree.header().setSectionResizeMode(1, QHeaderView.ResizeToContents)
    
    @Slot()
    def start_scan(self):
        if self.scanner_process and self.scanner_process.state() == QProcess.Running: return
        proxy_selection = self.scan_proxy_selector.currentText()
        self.start_scan_button.setEnabled(False); self.cancel_scan_button.setEnabled(True)
        self.scan_model.setRowCount(0); self.update_status("Starting scanner process...")
        self.scanner_buffer = ""
        self.scanner_process = QProcess()
        self.scanner_process.readyReadStandardOutput.connect(self.handle_scanner_output)
        self.scanner_process.readyReadStandardError.connect(self.handle_scanner_error)
        self.scanner_process.finished.connect(self.on_scanner_finished)
        self.scanner_process.errorOccurred.connect(self.on_scanner_process_error)
        args = ["--subnet", self.scan_subnet_input.text(), "--user", self.scan_user_input.text(), "--password", self.scan_pass_input.text()]
        if proxy_selection == "None (Direct Scan)": args.append("--no-proxy")
        else: 
            if not proxy_selection or proxy_selection not in self.config['proxies']:
                QMessageBox.warning(self, "Scanner Error", f"Proxy '{proxy_selection}' not found in settings.");
                self.start_scan_button.setEnabled(True); self.cancel_scan_button.setEnabled(False)
                return
            args.extend(["--proxy-port", str(self.config['proxies'][proxy_selection])])
        if self.quick_scan_checkbox.isChecked(): args.append("--quick-scan")
        python_executable = sys.executable
        script_path = os.path.join(os.path.dirname(os.path.realpath(__file__)), "scanner_process.py")
        self.scanner_process.start(python_executable, [script_path] + args)
    @Slot()
    def stop_scan(self):
        if self.scanner_process and self.scanner_process.state() == QProcess.Running: self.scanner_process.kill()
        else: self.start_scan_button.setEnabled(True); self.cancel_scan_button.setEnabled(False)
    @Slot()
    def handle_scanner_output(self):
        data = self.scanner_process.readAllStandardOutput().data().decode(errors='ignore'); self.scanner_buffer += data
        while '\n' in self.scanner_buffer:
            newline_pos = self.scanner_buffer.find('\n'); line = self.scanner_buffer[:newline_pos].strip()
            self.scanner_buffer = self.scanner_buffer[newline_pos+1:]
            if not line: continue
            if line.startswith("STATUS:"): self.update_status(line.replace("STATUS:", "").strip())
            elif line.startswith("RESULT:"):
                try:
                    result = json.loads(line.replace("RESULT:", "").strip()); host = result.get('host')
                    host_items = self.scan_model.findItems(host)
                    if not host_items: host_item = QStandardItem(host); host_item.setEditable(False); self.scan_model.appendRow(host_item)
                    else: host_item = host_items[0]
                    shares = result.get('shares', [])
                    for share in shares:
                        share_item = QStandardItem(share['name']); perms_item = QStandardItem(share['permissions'])
                        host_item.appendRow([QStandardItem(""), share_item, perms_item])
                except json.JSONDecodeError: print(f"[GUI ERROR] Failed to parse JSON result: {line}")
    @Slot()
    def handle_scanner_error(self):
        error_data = self.scanner_process.readAllStandardError().data().decode(errors='ignore'); print(f"[SCANNER STDERR]: {error_data.strip()}")
    @Slot(QProcess.ProcessError)
    def on_scanner_process_error(self, error):
        self.update_status("Error: Scanner process failed to start."); self.start_scan_button.setEnabled(True)
        self.cancel_scan_button.setEnabled(False); self.scanner_process = None
    @Slot()
    def on_scanner_finished(self):
        if self.scanner_buffer.strip(): self.handle_scanner_output()
        if self.scanner_process and self.scanner_process.exitStatus() == QProcess.NormalExit:
            final_message = f"Scan finished. Found {self.scan_model.rowCount()} host(s) with shares." if self.scan_model.rowCount() > 0 else "Scan finished. No accessible shares were found."
            self.update_status(final_message)
        else:
            if self.cancel_scan_button.isEnabled(): self.update_status("Scan cancelled or failed.")
        self.start_scan_button.setEnabled(True); self.cancel_scan_button.setEnabled(False); self.scanner_process = None
    @Slot()
    def start_preview(self):
        remote_path = self.get_selected_file_path()
        if remote_path and self.browser_worker: QMetaObject.invokeMethod(self.browser_worker, 'do_preview', Qt.QueuedConnection, Q_ARG(str, remote_path))
    @Slot()
    def start_download(self):
        remote_path = self.get_selected_file_path()
        if not (remote_path and self.browser_worker): return
        file_name = remote_path.split('/')[-1]; local_path, _ = QFileDialog.getSaveFileName(self, "Save File", file_name)
        if local_path: QMetaObject.invokeMethod(self.browser_worker, 'do_download', Qt.QueuedConnection, Q_ARG(str, remote_path), Q_ARG(str, local_path))
    @Slot()
    def disconnect(self):
        if self.browser_worker_thread and self.browser_worker_thread.isRunning():
            self.browser_worker.stop(); self.browser_worker_thread.quit(); self.browser_worker_thread.wait(3000)
        self.browser_worker = None; self.browser_worker_thread = None; self.reset_browser_ui()

    def reset_browser_ui(self):
        self.model.clear(); self.model.setHorizontalHeaderLabels(['Name', 'Size']); self.connect_button.setEnabled(True)
        self.disconnect_button.setEnabled(False); self.preview_button.setEnabled(False)
        self.download_button.setEnabled(False); self.search_button.setEnabled(False)
        self.clear_search_button.setEnabled(False); self.cancel_search_button.setVisible(False)
        self.search_button.setVisible(True); self.proxy_selector.setEnabled(True)
        self.update_status("Disconnected"); self.is_in_search_mode = False
        self.path_label.setText("Current Path: /")
        self.file_tree.header().setSectionResizeMode(0, QHeaderView.Stretch)
        self.file_tree.header().setSectionResizeMode(1, QHeaderView.ResizeToContents)

    @Slot(str)
    def update_status(self, message): self.statusBar().showMessage(f"Status: {message}")

    @Slot(list)
    def on_connection_success(self, file_list):
        self.update_status("Connected successfully!")
        self.disconnect_button.setEnabled(True)
        self.search_button.setEnabled(True)
        self.clear_search_button.setEnabled(False)
        self.browse_path('/')

        # --- MODIFIED: Automatically start caching if enabled in settings ---
        if self.browser_worker and self.config.get('auto_index', True):
            QMetaObject.invokeMethod(self.browser_worker, 'start_background_caching', Qt.QueuedConnection)

    @Slot(str)
    def on_connection_failed(self, error_message):
        QMessageBox.critical(self, "Connection Failed", error_message); self.reset_browser_ui()

    @Slot(str, bytes)
    def on_preview_ready(self, file_name, content):
        dialog = PreviewDialog(file_name, content, self); dialog.show()

    def get_selected_file_path(self):
        if self.tabs.currentWidget() != self.browser_tab: return None
        indexes = self.file_tree.selectionModel().selectedIndexes()
        if not indexes: return None
        item = self.model.itemFromIndex(indexes[0])
        item_text = item.text()
        if self.is_in_search_mode:
            return item_text if item_text and "No matching files found." not in item_text else None
        else:
            item_data = item.data(Qt.UserRole)
            if item_data != 'file': return None
            return f"{self.current_smb_path.rstrip('/')}/{item_text}" if self.current_smb_path != "/" else f"/{item_text}"

    @Slot()
    def on_selection_changed(self):
        path = self.get_selected_file_path()
        is_file = path is not None
        self.preview_button.setEnabled(is_file); self.download_button.setEnabled(is_file)

    def browse_path(self, path):
        self.is_in_search_mode = False; self.model.clear(); self.model.setHorizontalHeaderLabels(['Name', 'Size'])
        self.current_smb_path = path; self.path_label.setText(f"Current Path: {self.smb_share.text()}{path}")
        dir_icon = QApplication.style().standardIcon(QStyle.SP_DirIcon)
        try:
            files_and_dirs = self.browser_worker.browse_path(path)
            if path != "/":
                item = QStandardItem(dir_icon, ".."); item.setData("..", Qt.UserRole)
                size_item = QStandardItem(""); size_item.setEditable(False)
                self.model.appendRow([item, size_item])
            for entry in files_and_dirs:
                size_item = QStandardItem()
                if entry['type'] == 'dir':
                    icon = dir_icon
                    size_item.setText("")
                else:
                    icon = self.get_icon_for_filename(entry['name'])
                    size_str = self.format_file_size(entry['size'])
                    size_item.setText(size_str)
                    size_item.setTextAlignment(Qt.AlignRight | Qt.AlignVCenter)
                
                size_item.setEditable(False)
                item = QStandardItem(icon, entry['name']); item.setData(entry['type'], Qt.UserRole)
                item.setEditable(False)
                self.model.appendRow([item, size_item])
        except Exception as e:
           QMessageBox.warning(self, "Browse Error", f"Could not list path '{path}'.\n\n{e}")
        self.file_tree.header().setSectionResizeMode(0, QHeaderView.Stretch)
        self.file_tree.header().setSectionResizeMode(1, QHeaderView.ResizeToContents)


    def on_item_double_clicked(self, index):
        if self.is_in_search_mode: self.start_preview(); return
        if not (self.browser_worker and self.browser_worker.smb_connection): return
        item = self.model.itemFromIndex(index); item_type = item.data(Qt.UserRole); item_text = item.text()
        new_path = ""
        if item_type == "..":
            if self.current_smb_path != "/":
                parts = self.current_smb_path.strip('/').split('/'); new_path = "/" + "/".join(parts[:-1]) if len(parts) > 1 else "/"
        elif item_type == "dir":
            new_path = f"{self.current_smb_path.rstrip('/')}/{item_text}" if self.current_smb_path != "/" else f"/{item_text}"
        else: self.start_preview(); return
        self.browse_path(new_path)

    def show_scanner_context_menu(self, pos: QPoint):
        index = self.scan_results_tree.indexAt(pos)
        if not index.isValid(): return
        item = self.scan_model.itemFromIndex(index)
        if not item.parent(): return
        menu = QMenu(); open_action = menu.addAction("Open in Browser")
        action = menu.exec(self.scan_results_tree.mapToGlobal(pos))
        if action == open_action: self.open_share_in_browser()

    @Slot()
    def on_scanner_selection_changed(self):
        indexes = self.scan_results_tree.selectionModel().selectedIndexes()
        is_share = False
        if indexes:
            item = self.scan_model.itemFromIndex(indexes[0])
            if item.parent(): is_share = True
        self.open_share_button.setEnabled(is_share)

    @Slot()
    def open_share_in_browser(self):
        indexes = self.scan_results_tree.selectionModel().selectedIndexes()
        if not indexes: return
        item = self.scan_model.itemFromIndex(indexes[0])
        if item.parent():
            host_item = item.parent(); share_item = host_item.child(item.row(), 1)
            host = host_item.text(); share_name = share_item.text()
            proxy_selection = self.scan_proxy_selector.currentText()
            self.proxy_selector.setCurrentText(proxy_selection)
            self.smb_host.setText(host); self.smb_share.setText(share_name)
            self.smb_user.setText(self.scan_user_input.text()); self.smb_pass.setText(self.scan_pass_input.text())
            self.tabs.setCurrentWidget(self.browser_tab)
            self.update_status(f"Loaded {host}/{share_name} into browser. Click Connect.")

    def closeEvent(self, event):
        self.save_config()
        self.disconnect()
        self.stop_scan()
        event.accept()

if __name__ == "__main__":
    app = QApplication(sys.argv)
    window = SMBBrowserApp()
    window.show()
    sys.exit(app.exec())
